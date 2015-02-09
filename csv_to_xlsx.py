from openpyxl import Workbook
from openpyxl.cell import get_column_letter
import openpyxl
import delimited_file_utils
import os
from openpyxl.cell import get_column_letter

red = openpyxl.styles.colors.RED
blue = openpyxl.styles.colors.BLUE
mygray = openpyxl.styles.colors.Color('D0D0D0')
lg_fill = openpyxl.styles.fills.PatternFill(fill_type='solid', \
                                            fgColor=mygray, \
                                            bgColor=blue)
lg_style = openpyxl.styles.Style(fill=lg_fill)


def get_max_width(col_in):
    """Find the maximum number of characters for any string in a
    column"""
    max_width = -1

    for item in col_in:
        cur_width = len(item)
        if cur_width > max_width:
            max_width = cur_width

    return max_width


def csv_to_xlsx(pathin, pathout=None, **kwargs):
    if pathout is None:
        fno, ext = os.path.splitext(pathin)
        pathout = fno + '.xlsx'
        
    array = delimited_file_utils.open_delimited_with_sniffer_and_check(pathin)
    array_to_xlsx(array, pathout, **kwargs)


def array_to_xlsx(arrayin, filename, alt_gray=True, title='Sheet1', \
                  set_max_width=False):
    wb = Workbook()
    dest_filename = filename

    ws = wb.worksheets[0]
    ws.title = title
    
    for row_index, row in enumerate(arrayin):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            cur_cell = ws.cell('%s%s'%(column_letter, (row_index + 1)))
            try:
                myfloat = float(cell)
                cur_cell.value = myfloat
                #cur_cell.number_format = '0.00'
                #print('yeah, numbers')
            except:
                cur_cell.value = cell

            if alt_gray:
                if row_index % 2:
                    cur_cell.style = lg_style


    if set_max_width:
        nr, nc = arrayin.shape
        for i in range(nc):
            max_width = get_max_width(arrayin[:,i])
            ws.column_dimensions[get_column_letter(i+1)].width = max_width


    wb.save(filename)
